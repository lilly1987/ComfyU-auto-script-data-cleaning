# -*- coding: utf-8 -*-
"""
Read PNG metadata from files inside W:\\fail.
"""
from __future__ import annotations

import json
import os
import sqlite3
import struct
import sys
import zlib
from datetime import datetime
from pathlib import Path
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


PNG_SIGNATURE = b"\x89PNG\r\n\x1a\n"
TARGET_DIR = Path(r"W:\fail")
DB_PATH = Path(__file__).resolve().parent.parent / "error.db"
EXCEL_PATH = DB_PATH.with_suffix(".xlsx")
TABLE_NAME = "LoraLoader"
SCHEMA_COLUMNS = [
    ("id", "INTEGER PRIMARY KEY AUTOINCREMENT"),
    ("recorded_at", "TEXT"),
    ("ckpt_name", "TEXT"),
    ("lora_name", "TEXT"),
    ("steps", "INTEGER"),
    ("cfg", "REAL"),
    ("sampler_name", "TEXT"),
    ("scheduler", "TEXT"),
    ("denoise", "REAL"),
    ("positive", "TEXT"),
    ("negative", "TEXT"),
    ("strength_model", "REAL"),
    ("strength_clip", "REAL"),
    ("A", "REAL"),
    ("B", "REAL"),
    ("block_vector", "TEXT"),
]
DATA_COLUMNS = [name for name, _ in SCHEMA_COLUMNS if name != "id"]


def _decode_text(data: bytes) -> str:
    for encoding in ("utf-8", "latin-1"):
        try:
            return data.decode(encoding)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def _parse_text_chunk(chunk_type: bytes, data: bytes) -> tuple[str, str]:
    if chunk_type == b"tEXt":
        keyword, text = data.split(b"\x00", 1)
        return _decode_text(keyword), _decode_text(text)

    if chunk_type == b"zTXt":
        keyword, rest = data.split(b"\x00", 1)
        compression_method = rest[0]
        compressed_text = rest[1:]
        if compression_method != 0:
            raise ValueError(f"Unsupported zTXt compression method: {compression_method}")
        text = zlib.decompress(compressed_text)
        return _decode_text(keyword), _decode_text(text)

    if chunk_type == b"iTXt":
        keyword, rest = data.split(b"\x00", 1)
        compression_flag = rest[0]
        compression_method = rest[1]
        rest = rest[2:]

        _, rest = rest.split(b"\x00", 1)  # language tag
        _, text = rest.split(b"\x00", 1)  # translated keyword

        if compression_flag:
            if compression_method != 0:
                raise ValueError(f"Unsupported iTXt compression method: {compression_method}")
            text = zlib.decompress(text)

        return _decode_text(keyword), _decode_text(text)

    raise ValueError(f"Unsupported chunk type: {chunk_type!r}")


def read_png_metadata(file_path: Path) -> Dict[str, List[str]]:
    metadata: Dict[str, List[str]] = {}

    with file_path.open("rb") as handle:
        signature = handle.read(len(PNG_SIGNATURE))
        if signature != PNG_SIGNATURE:
            raise ValueError("Not a valid PNG file")

        while True:
            raw_length = handle.read(4)
            if not raw_length:
                break
            if len(raw_length) != 4:
                raise ValueError("Unexpected end of file while reading chunk length")

            length = struct.unpack(">I", raw_length)[0]
            chunk_type = handle.read(4)
            chunk_data = handle.read(length)
            handle.read(4)  # CRC

            if len(chunk_type) != 4 or len(chunk_data) != length:
                raise ValueError("Unexpected end of file while reading chunk data")

            if chunk_type in {b"tEXt", b"zTXt", b"iTXt"}:
                key, value = _parse_text_chunk(chunk_type, chunk_data)
                metadata.setdefault(key, []).append(value)

            if chunk_type == b"IEND":
                break

    return metadata


def ensure_database(connection: sqlite3.Connection) -> None:
    connection.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {TABLE_NAME} (
            {", ".join(f"{name} {column_type}" for name, column_type in SCHEMA_COLUMNS)}
        )
        """
    )
    connection.commit()
    migrate_database(connection)


def get_table_columns(connection: sqlite3.Connection, table_name: str) -> List[str]:
    rows = connection.execute(f"PRAGMA table_info({table_name})").fetchall()
    return [row[1] for row in rows]


def migrate_database(connection: sqlite3.Connection) -> None:
    table_columns = get_table_columns(connection, TABLE_NAME)

    if "lora_key" in table_columns and "lora_name" not in table_columns:
        connection.execute(f"ALTER TABLE {TABLE_NAME} RENAME COLUMN lora_key TO lora_name")
        connection.commit()
        table_columns = get_table_columns(connection, TABLE_NAME)

    for column_name, column_type in SCHEMA_COLUMNS:
        if column_name not in table_columns:
            connection.execute(f"ALTER TABLE {TABLE_NAME} ADD COLUMN {column_name} {column_type}")
            connection.commit()
            table_columns = get_table_columns(connection, TABLE_NAME)

    desired_order = [name for name, _ in SCHEMA_COLUMNS]
    if table_columns != desired_order:
        temp_table = f"{TABLE_NAME}_new"
        connection.execute(
            f"""
            CREATE TABLE {temp_table} (
                {", ".join(f"{name} {column_type}" for name, column_type in SCHEMA_COLUMNS)}
            )
            """
        )

        available_columns = [column for column in desired_order if column in table_columns and column != "id"]
        select_sql = ", ".join(available_columns)
        insert_sql = ", ".join(available_columns)
        connection.execute(
            f"""
            INSERT INTO {temp_table} ({insert_sql})
            SELECT {select_sql}
            FROM {TABLE_NAME}
            """
        )
        connection.execute(f"DROP TABLE {TABLE_NAME}")
        connection.execute(f"ALTER TABLE {temp_table} RENAME TO {TABLE_NAME}")
        connection.commit()


def extract_lora_records(file_path: Path, metadata: Dict[str, List[str]]) -> List[Dict[str, object]]:
    prompt_values = metadata.get("prompt", [])
    if not prompt_values:
        return []

    prompt_data = json.loads(prompt_values[0])
    recorded_at = datetime.now().isoformat(timespec="seconds")
    records: List[Dict[str, object]] = []
    ckpt_name = None
    steps = None
    cfg = None
    sampler_name = None
    scheduler = None
    denoise = None
    positive = None
    negative = None

    checkpoint_loader = prompt_data.get("CheckpointLoaderSimple", {})
    checkpoint_inputs = checkpoint_loader.get("inputs", {})
    checkpoint_path = checkpoint_inputs.get("ckpt_name")
    if checkpoint_path:
        ckpt_name = Path(str(checkpoint_path)).stem

    ksampler = prompt_data.get("KSampler", {})
    ksampler_inputs = ksampler.get("inputs", {})
    steps = ksampler_inputs.get("steps")
    cfg = ksampler_inputs.get("cfg")
    sampler_name = ksampler_inputs.get("sampler_name")
    scheduler = ksampler_inputs.get("scheduler")
    denoise = ksampler_inputs.get("denoise")

    positive_wildcard = prompt_data.get("positiveWildcard", {})
    positive_inputs = positive_wildcard.get("inputs", {})
    positive = positive_inputs.get("populated_text")

    negative_wildcard = prompt_data.get("negativeWildcard", {})
    negative_inputs = negative_wildcard.get("inputs", {})
    negative = negative_inputs.get("populated_text")

    for node_key, node_value in sorted(prompt_data.items()):
        if not node_key.startswith("LoraLoader"):
            continue

        inputs = node_value.get("inputs", {})
        lora_name = inputs.get("lora_name")
        if not lora_name:
            continue

        lora_key = Path(str(lora_name)).stem
        records.append(
            {
                "png_file": file_path.name,
                "lora_name": lora_key,
                "ckpt_name": ckpt_name,
                "steps": steps,
                "cfg": cfg,
                "sampler_name": sampler_name,
                "scheduler": scheduler,
                "denoise": denoise,
                "positive": positive,
                "negative": negative,
                "strength_model": inputs.get("strength_model"),
                "strength_clip": inputs.get("strength_clip"),
                "A": inputs.get("A"),
                "B": inputs.get("B"),
                "block_vector": inputs.get("block_vector"),
                "recorded_at": recorded_at,
            }
        )

    return records


def insert_lora_records(connection: sqlite3.Connection, records: List[Dict[str, object]]) -> None:
    if not records:
        return

    table_columns = get_table_columns(connection, TABLE_NAME)
    normalized_records: List[Dict[str, object]] = []

    for record in records:
        normalized_record = dict(record)
        if "lora_key" in table_columns and "lora_name" not in table_columns:
            normalized_record["lora_key"] = normalized_record.get("lora_name")
        normalized_records.append(normalized_record)

    insert_columns = [
        column
        for column in DATA_COLUMNS
        if column in table_columns
    ]
    column_sql = ",\n            ".join(insert_columns)
    value_sql = ",\n            ".join(f":{column}" for column in insert_columns)

    connection.executemany(
        f"""
        INSERT INTO {TABLE_NAME} (
            {column_sql}
        ) VALUES (
            {value_sql}
        )
        """,
        normalized_records,
    )
    connection.commit()


def print_lora_summary(records: List[Dict[str, object]]) -> None:
    if not records:
        print("No LoraLoader records found in prompt metadata.")
        return

    first_record = records[0]
    print(f"ckpt_name: {first_record.get('ckpt_name')}")
    print(
        "sampler: "
        f"steps={first_record.get('steps')}, "
        f"cfg={first_record.get('cfg')}, "
        f"sampler_name={first_record.get('sampler_name')}, "
        f"scheduler={first_record.get('scheduler')}, "
        f"denoise={first_record.get('denoise')}"
    )
    print(f"positive length: {len(first_record.get('positive') or '')}")
    print(f"negative length: {len(first_record.get('negative') or '')}")
    print("loras:")
    for record in records:
        print(
            f"  - {record.get('lora_name')}: "
            f"strength_model={record.get('strength_model')}, "
            f"strength_clip={record.get('strength_clip')}, "
            f"A={record.get('A')}, "
            f"B={record.get('B')}"
        )


def vacuum_database(connection: sqlite3.Connection) -> None:
    connection.execute("VACUUM")
    connection.commit()


def export_to_excel(connection: sqlite3.Connection, output_path: Path) -> None:
    query = f"""
    SELECT {", ".join(DATA_COLUMNS)}
    FROM {TABLE_NAME}
    ORDER BY id
    """
    rows = connection.execute(query).fetchall()

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = TABLE_NAME

    worksheet.append(DATA_COLUMNS)
    for row in rows:
        worksheet.append(list(row))

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False

    wrap_columns = {"I", "J", "O"}  # positive, negative, block_vector
    for column_cells in worksheet.iter_cols(min_row=2, max_row=worksheet.max_row):
        column_letter = column_cells[0].column_letter
        max_length = len(str(worksheet[f"{column_letter}1"].value or ""))
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, min(len(value), 80))
            if column_letter in wrap_columns:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 60)

    worksheet.row_dimensions[1].height = 22

    if worksheet.max_row >= 2 and worksheet.max_column >= 1:
        table_ref = f"A1:{worksheet.cell(row=worksheet.max_row, column=worksheet.max_column).coordinate}"
        table = Table(displayName=TABLE_NAME, ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    workbook.save(output_path)


def delete_processed_png(file_path: Path) -> None:
    try:
        os.chmod(file_path, 0o666)
    except OSError:
        pass
    file_path.unlink()


def main() -> int:
    if not TARGET_DIR.exists():
        print(f"[ERROR] Folder not found: {TARGET_DIR}")
        return 1

    png_files = sorted(TARGET_DIR.glob("*.png"))
    print("=" * 80)
    print(f"PNG metadata scan: {TARGET_DIR}")
    print(f"PNG files found: {len(png_files)}")
    print("=" * 80)

    error_count = 0
    total_lora_records = 0

    connection = sqlite3.connect(DB_PATH)
    ensure_database(connection)

    try:
        if not png_files:
            print("No PNG files found.")
        else:
            for index, file_path in enumerate(png_files, start=1):
                print()
                print("-" * 80)
                print(f"[{index}/{len(png_files)}] {file_path.name}")
                print("-" * 80)

                try:
                    metadata = read_png_metadata(file_path)
                except Exception as exc:
                    error_count += 1
                    print(f"[ERROR] Failed to read metadata: {exc}")
                    continue

                if not metadata:
                    print("No PNG text metadata found.")
                    continue

                try:
                    lora_records = extract_lora_records(file_path, metadata)
                    print_lora_summary(lora_records)
                    insert_lora_records(connection, lora_records)
                    total_lora_records += len(lora_records)
                    print(f"LoraLoader rows inserted: {len(lora_records)}")
                except Exception as exc:
                    error_count += 1
                    print(f"[ERROR] Failed to write LoraLoader history: {exc}")
                    continue

                try:
                    delete_processed_png(file_path)
                    print("PNG deleted after successful processing.")
                except Exception as exc:
                    error_count += 1
                    print(f"[ERROR] Failed to delete PNG: {exc}")

        vacuum_database(connection)
        export_to_excel(connection, EXCEL_PATH)
        print(f"Excel exported: {EXCEL_PATH}")
    except Exception as exc:
        error_count += 1
        print(f"[ERROR] Finalization failed: {exc}")
    finally:
        connection.close()

    print("=" * 80)
    print(f"Done. Errors: {error_count}, LoraLoader rows inserted: {total_lora_records}")
    print(f"Database: {DB_PATH}")
    print(f"Excel: {EXCEL_PATH}")
    print("=" * 80)
    return 0 if error_count == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
